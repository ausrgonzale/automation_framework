import glob
import os
import subprocess

import openpyxl
import pytest


def read_excel_as_list_of_dicts(path):
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        raise AssertionError(f"Could not open Excel file '{path}': {e}")
    ws = wb.active
    if ws is None:
        raise AssertionError(f"No worksheet found in Excel file '{path}'")
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def compare_excels(generated_path, expected_path):
    gen_rows = read_excel_as_list_of_dicts(generated_path)
    exp_rows = read_excel_as_list_of_dicts(expected_path)
    assert len(gen_rows) == len(
        exp_rows
    ), f"Row count mismatch: {len(gen_rows)} != {len(exp_rows)}"
    for i, (g, e) in enumerate(zip(gen_rows, exp_rows)):
        for key in e:
            assert g.get(key) == e.get(
                key
            ), f"Mismatch at row {i+2}, column '{key}': {g.get(key)} != {e.get(key)}"


@pytest.mark.parametrize(
    "generated,expected",
    [
        (
            "testdata/userstories/excel/generated_testcases_latest.xlsx",
            "testdata/userstories/excel/expected_testcases.xlsx",
        ),
        (
            "testdata/userstories/jira/generated_testcases_jira_latest.xlsx",
            "testdata/userstories/jira/expected_testcases_jira.xlsx",
        ),
    ],
)
def test_excel_output_matches_expected(generated, expected):
    # Run the script to generate Excel files before assertions
    subprocess.run(["python", "scripts/orchestrate_test_generation.py"], check=False)

    # Use glob to find the latest generated file matching the pattern
    if "generated_testcases_latest.xlsx" in generated:
        pattern = generated.replace("_latest", "_*")
        matches = glob.glob(pattern)
        assert matches, f"No generated Excel files found matching pattern: {pattern}"
        # Pick the most recent file
        generated_file = max(matches, key=os.path.getmtime)
    else:
        generated_file = generated

    assert os.path.exists(generated_file), f"Generated file not found: {generated_file}"
    assert os.path.exists(expected), f"Expected file not found: {expected}"
    compare_excels(generated_file, expected)
