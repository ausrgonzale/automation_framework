"""
===============================================================================
File: excel_storage.py

Location:
    repositories/

Component Type:
    Storage Engine

Purpose:
    Provides low-level Excel file operations.

Design Principles:

    - Deterministic behavior
    - Safe initialization
    - Explicit error handling
    - Structured logging
    - Repository-friendly interface

===============================================================================
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)


class ExcelStorage:
    """
    Provides low-level Excel file operations.
    """

    # ---------------------------------------------------------------------
    # Initialization
    # ---------------------------------------------------------------------

    def __init__(self, filename: str) -> None:

        self.filename: str = str(filename)
        self.file_path: Path = Path(self.filename)

        self.workbook: Optional[Workbook] = None
        self.sheet: Optional[Worksheet] = None

        self._ensure_directory()

    # ---------------------------------------------------------------------
    # Directory Safety
    # ---------------------------------------------------------------------

    def _ensure_directory(self) -> None:

        try:

            directory = self.file_path.parent

            if not directory.exists():

                directory.mkdir(
                    parents=True,
                    exist_ok=True,
                )

                logger.info(
                    "Created directory for Excel file",
                    extra={
                        "directory": str(directory),
                    },
                )

        except Exception:

            logger.exception(
                "Failed to create directory",
                extra={
                    "directory": str(self.file_path.parent),
                },
            )

            raise

    # ---------------------------------------------------------------------
    # Internal Guards
    # ---------------------------------------------------------------------

    def _require_workbook(self) -> Workbook:

        if self.workbook is None:
            raise RuntimeError("Workbook is not loaded")

        return self.workbook

    def _require_sheet(self) -> Worksheet:

        if self.sheet is None:
            raise RuntimeError("Worksheet is not loaded")

        return self.sheet

    # ---------------------------------------------------------------------
    # File Lifecycle
    # ---------------------------------------------------------------------

    def create_file(
        self,
        sheet_name: str = "Sheet1",
        headers: Optional[List[str]] = None,
    ) -> None:

        try:

            logger.info(
                "Creating Excel file",
                extra={
                    "file": self.filename,
                    "sheet": sheet_name,
                },
            )

            self.workbook = Workbook()

            active_sheet = self.workbook.active

            if active_sheet is None:
                raise RuntimeError(
                    "Failed to obtain active worksheet"
                )

            self.sheet = active_sheet
            self.sheet.title = sheet_name

            if headers:

                for col_num, header in enumerate(
                    headers,
                    start=1,
                ):

                    self.sheet.cell(
                        row=1,
                        column=col_num,
                        value=header,
                    )

            workbook = self._require_workbook()

            workbook.save(self.filename)

            logger.info(
                "Excel file created successfully",
                extra={
                    "file": self.filename,
                    "sheet": sheet_name,
                },
            )

        except Exception:

            logger.exception(
                "Failed to create Excel file",
                extra={
                    "file": self.filename,
                    "sheet": sheet_name,
                },
            )

            raise

    # ---------------------------------------------------------------------

    def load_file(
        self,
        auto_create: bool = True,
        sheet_name: str = "Sheet1",
    ) -> bool:

        try:

            if not self.file_path.exists():

                logger.info(
                    "Excel file does not exist",
                    extra={
                        "file": self.filename
                    },
                )

                if auto_create:

                    self.create_file(
                        sheet_name=sheet_name,
                        headers=[
                            "id",
                            "title",
                            "steps",
                            "expected_result",
                            "priority",
                        ],
                    )

                else:

                    return False

            self.workbook = load_workbook(
                self.filename
            )

            self.sheet = self.workbook.active

            logger.info(
                "Excel file loaded successfully",
                extra={
                    "file": self.filename
                },
            )

            return True

        except Exception:

            logger.exception(
                "Failed to load Excel file",
                extra={
                    "file": self.filename
                },
            )

            raise

    # ---------------------------------------------------------------------
    # Row Operations
    # ---------------------------------------------------------------------

    def create_row(
        self,
        values: List[Any],
    ) -> None:

        try:

            sheet = self._require_sheet()

            sheet.append(values)

            workbook = self._require_workbook()

            workbook.save(self.filename)

            logger.info(
                "Row added to Excel",
                extra={
                    "file": self.filename,
                    "values_count": len(values),
                },
            )

        except Exception:

            logger.exception(
                "Failed to create row",
                extra={
                    "file": self.filename
                },
            )

            raise

    # ---------------------------------------------------------------------
    # Read Operations
    # ---------------------------------------------------------------------

    def read_all(self) -> List[List[Any]]:
        """
        Read all rows from the active worksheet.
        """

        try:

            sheet = self._require_sheet()

            rows: List[List[Any]] = []

            for row in sheet.iter_rows(values_only=True):

                rows.append(list(row))

            logger.info(
                "Read all rows from Excel",
                extra={
                    "file": self.filename,
                    "row_count": len(rows),
                },
            )

            return rows

        except Exception:

            logger.exception(
                "Failed to read rows",
                extra={
                    "file": self.filename
                },
            )

            raise

    # ---------------------------------------------------------------------
    # Metadata
    # ---------------------------------------------------------------------

    def get_info(self) -> Dict[str, Any]:
        """
        Retrieve workbook metadata.
        """

        try:

            workbook = self._require_workbook()

            sheet_names = workbook.sheetnames

            info: Dict[str, Any] = {
                "file": self.filename,
                "all_sheets": sheet_names,
                "sheet_count": len(sheet_names),
            }

            logger.info(
                "Retrieved workbook info",
                extra=info,
            )

            return info

        except Exception:

            logger.exception(
                "Failed to retrieve workbook info",
                extra={
                    "file": self.filename
                },
            )

            raise